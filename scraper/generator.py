from openpyxl import Workbook

# Generator class to create xlsx output.

# !--- output path ---! #
filename = 'casedata.xlsx'
dest = '~/Downloads/' + filename
#dest = '../exceltests/test.xlsx'

# Write data headers to xlsx
def write_headers(worksheet):
    headers = ['Case Number', 'Attorney Name', 'Address', 'Party Type', 'City/St/Zip', 'Party no.', 'Name', 'Party Type', 'Address', 'City', 'State/Zip', 'Party no.', 'Name', 'Party Type', 'Address', 'City', 'State/Zip', 'Party no.', 'Name', 'Party Type', 'Address', 'City', 'State/Zip']
    row = 1
    col = 1
    for header in headers:
        worksheet.cell(row, col, value=header)
        col+=1

# Write scraped data to xlsx
def generate(workbookArgs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Case Data"
    write_headers(ws)
    row = 2
    col = 1
    for case_row in workbookArgs:
        for data in case_row:
            ws.cell(row, col, value=data)
            col+=1
        row+=1
        col=1
    wb.save(dest)
    print('File ' + filename + ' saved at ', dest)


